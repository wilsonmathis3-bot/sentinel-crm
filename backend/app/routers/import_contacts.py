"""Contact import router — CSV/XLSX with fuzzy column mapping, dedupe, dry-run."""
import csv
import io
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.database import get_db
from app import crud, schemas, models

router = APIRouter()

# ---------------------------------------------------------------------------
# Fuzzy column mapping
# ---------------------------------------------------------------------------
COLUMN_ALIASES = {
    "first_name": [
        "first name", "firstname", "first_name", "fname",
        "given name", "given_name", "name first", "name_first"
    ],
    "last_name": [
        "last name", "lastname", "last_name", "lname",
        "surname", "family name", "family_name", "name last", "name_last"
    ],
    "email": [
        "email", "e-mail", "email address", "e-mail address",
        "email_address", "mail", "e mail"
    ],
    "phone": [
        "phone", "telephone", "mobile", "cell", "phone number",
        "phone_number", "tel", "cellphone", "cell phone"
    ],
    "company": [
        "company", "organization", "org", "company name",
        "company_name", "organisation", "employer", "firm"
    ],
    "city": [
        "city", "town", "municipality"
    ],
    "state": [
        "state", "province", "region", "territory", "county"
    ],
    "industry": [
        "industry", "sector", "vertical", "business type", "business_type"
    ],
    "notes": [
        "notes", "comments", "remarks", "note", "description",
        "memo", "additional info", "additional_info"
    ],
}


def _normalize_header(h: str) -> str:
    return h.strip().lower().replace("_", " ").replace("-", " ")


def _build_column_map(headers: List[str]) -> dict:
    """Return {canonical_field: header_index} using fuzzy matching."""
    normalized = [_normalize_header(h) for h in headers]
    mapping = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            norm_alias = _normalize_header(alias)
            for idx, nh in enumerate(normalized):
                if nh == norm_alias or nh.replace(" ", "") == norm_alias.replace(" ", ""):
                    mapping[canonical] = idx
                    break
            if canonical in mapping:
                break
    return mapping


# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------
class RowError(BaseModel):
    row_number: int
    raw_data: dict
    errors: List[str]


class ImportSummary(BaseModel):
    imported: int
    skipped_duplicates: int
    errors: int
    dry_run: bool
    column_map: dict
    row_errors: List[RowError]


# ---------------------------------------------------------------------------
# File parsing
# ---------------------------------------------------------------------------
def _parse_csv(content: bytes) -> tuple:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    rows = list(reader)
    return headers, rows


def _parse_xlsx(content: bytes) -> tuple:
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="Excel file has no active worksheet")

    headers = [cell.value for cell in ws[1]]
    headers = [str(h).strip() if h is not None else "" for h in headers]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for idx, header in enumerate(headers):
            row_dict[header] = row[idx] if idx < len(row) else None
        rows.append(row_dict)
    return headers, rows


def _parse_file(file: UploadFile) -> tuple:
    content = file.file.read()
    filename = (file.filename or "").lower()
    if filename.endswith(".csv"):
        return _parse_csv(content)
    elif filename.endswith((".xlsx", ".xls")):
        return _parse_xlsx(content)
    else:
        raise HTTPException(status_code=400, detail="File must be .csv or .xlsx")


# ---------------------------------------------------------------------------
# Row validation & transformation
# ---------------------------------------------------------------------------
def _validate_row(row: dict, col_map: dict, row_number: int) -> tuple:
    """Return (contact_data: dict | None, errors: list)."""
    errors = []
    data = {}

    # Required fields
    first_name_idx = col_map.get("first_name")
    last_name_idx = col_map.get("last_name")
    email_idx = col_map.get("email")

    if first_name_idx is None:
        errors.append("First Name column not found")
    else:
        val = str(row.get(list(row.keys())[first_name_idx], "")).strip()
        if not val:
            errors.append("First Name is required")
        data["first_name"] = val

    if last_name_idx is None:
        errors.append("Last Name column not found")
    else:
        val = str(row.get(list(row.keys())[last_name_idx], "")).strip()
        if not val:
            errors.append("Last Name is required")
        data["last_name"] = val

    if email_idx is None:
        errors.append("Email column not found")
    else:
        val = str(row.get(list(row.keys())[email_idx], "")).strip()
        if not val:
            errors.append("Email is required")
        elif "@" not in val or "." not in val.split("@")[-1]:
            errors.append(f"Invalid email: {val}")
        data["email"] = val.lower()

    # Optional fields
    for field in ["phone", "company", "city", "state", "industry", "notes"]:
        idx = col_map.get(field)
        if idx is not None:
            val = row.get(list(row.keys())[idx])
            data[field] = str(val).strip() if val is not None else None

    if errors:
        return None, errors
    return data, []


# ---------------------------------------------------------------------------
# Import endpoint
# ---------------------------------------------------------------------------
@router.post("/import", response_model=ImportSummary)
async def import_contacts(
    file: UploadFile = File(...),
    dry_run: bool = Form(False),
    db: Session = Depends(get_db)
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file uploaded")

    headers, raw_rows = _parse_file(file)
    if not headers:
        raise HTTPException(status_code=400, detail="File has no headers")

    col_map = _build_column_map(headers)
    if "email" not in col_map:
        raise HTTPException(
            status_code=400,
            detail=f"Could not find an email column. Headers detected: {headers}"
        )

    imported = 0
    skipped_duplicates = 0
    row_errors: List[RowError] = []

    # Pre-load existing emails for dedupe (case-insensitive)
    existing_emails = set()
    for contact in db.query(models.Contact).all():
        if contact.email:
            existing_emails.add(contact.email.lower())

    for idx, raw_row in enumerate(raw_rows, start=2):  # row 1 is header
        contact_data, errors = _validate_row(raw_row, col_map, idx)
        if errors:
            row_errors.append(RowError(row_number=idx, raw_data=raw_row, errors=errors))
            continue

        email = contact_data["email"]
        if email in existing_emails:
            skipped_duplicates += 1
            continue

        if not dry_run:
            try:
                db_contact = crud.create_contact(db, schemas.ContactCreate(**contact_data))
                existing_emails.add(email)
                imported += 1
            except Exception as exc:
                row_errors.append(RowError(row_number=idx, raw_data=raw_row, errors=[str(exc)]))
        else:
            imported += 1

    return ImportSummary(
        imported=imported,
        skipped_duplicates=skipped_duplicates,
        errors=len(row_errors),
        dry_run=dry_run,
        column_map={k: headers[v] for k, v in col_map.items()},
        row_errors=row_errors
    )
